from table import Table
from openpyxl import load_workbook

import csv
import io
from openai import OpenAI

class TableHelper:
    HEADER_COLOR_INDEX = 4

    def __init__(self):
        self.max_blank_column = 1
        self.max_blank_row = 1
        self.max_row = 100
        self.max_columns = 20

        self.tables = []

    def is_my_digit(self, cell_value):
        return cell_value.replace(',', '').replace('$', '').strip().isdigit()

    def is_my_string(self, cell_value):
        return type(cell_value) == str

    def is_table_header(self, cell):
        return cell.fill.start_color.index == self.HEADER_COLOR_INDEX

    def is_gray_area(self, cell):
        return cell.fill.start_color.index == 'FF7F7F7F'

    def open_file(self, path):
        workbook = load_workbook(filename = path, data_only = True)
        sheet = workbook.active
        return sheet

    # rules :
    # 1. left side could have no blank cells
    # 2. left side could have 2 blank cells, this cell is colored,
    # 3  left side could have 1 blank cell,  then left of left cell is not blank, this cell is colored.
    # 4. colored cell could have 1 blank cell at right side, then the next right side cell is colored.
    # 5. the next down cell could be colored.
    def is_table_top_left(self, sheet, i, j):
        rows = list(sheet.iter_rows())
        #print(f"== cell: {rows[i][j]}, value: {rows[i][j].value}")
        if (j == 0 and self.is_table_header(rows[i][j]) ) or \
                (j == 1 and self.is_table_header(rows[i][j]) and  rows[i][j - 1].value == None)  or \
                (j >= 2 and self.is_table_header(rows[i][j]) and rows[i][j - 2].value == None and rows[i][j - 1].value == None) or \
                (j >= 2 and self.is_table_header(rows[i][j]) and (rows[i][j - 2].value != None and not self.is_table_header(rows[i][j - 2]) ) and rows[i][j - 1].value == None):

            # if left cell is colored, this cell is not  ( 对应example 0 G51, F51)
            if j >= 2 and self.is_table_header(rows[i][j-1]):
                return False

            # is the above cell is header? if so, this cell is not header
            if i == 0:
                return True
            else:
                return not self.is_table_header(rows[i-1][j])

        else:
            return False

    # rule: if the right two cells are blank, it's right border of this table
    def is_table_top_right(self, sheet, i, j):
        rows = list(sheet.iter_rows())
        row = rows[i]

        return self.is_table_header(row[j]) and \
                (not self.is_table_header(row[j + 1])) and \
                (not self.is_table_header(row[j + 2]))

    # the table's bottom left,should follow these rules:
    # 1. must be string, not pure number
    # 2. checking from current cell to downside, could have 1 blank row. if you found 2 blank rows, it's found
    # 3. should not appear new colored cell.
    #
    def is_table_bottom_left(self, sheet, i, j):
        rows = list(sheet.iter_rows())

        #print(f"== checking: {rows[i][j]}, value: {rows[i][j].value}, +1: {rows[i+1][j].value}, +2: {rows[i+2][j].value}")

        # (case1: 本身是str, 下面有2个空白行，一定是结束)
        # case1: it's a str type var, with 2 blank(or 1 blank 1 zero) at the down rows
        if self.is_my_string(rows[i][j].value)  and \
                (rows[i+1][j].value == None) and \
                (rows[i+2][j].value == None or rows[i+2][j].value == 0) :
            return True

        # case2: it's a str type var, and with a blank row (to down) , then with a colored row.
        if self.is_my_string(rows[i][j].value)  and \
                (rows[i+1][j].value == None and self.is_table_header(rows[i+2][j])):
            return True

        return False


    # get table's top, left, right , bottom
    def get_tables(self, sheet):

        # get all tables top left
        for i, row in enumerate(sheet.iter_rows()):
            for j, cell in enumerate(row):
                if i > self.max_row:
                    break
                #print(f"cell: {cell.value}, i: {i}, j: {j}, {row[j].value}")
                if(self.is_table_top_left(sheet, i, j)):
                    table = Table()
                    print(f"=== found table top left: column: {cell.column}, row: {cell.row}, value: {cell.value}")
                    table.name = cell.value
                    table.left = cell.column - 1
                    table.top = cell.row - 1
                    self.tables.append(table)

        for table in self.tables:
            #print(f"==checking for table: {table}")
            # get all tables top right , 只考察一行即可( table.top)
            target_row = list(sheet.iter_rows())[table.top]
            for cell in target_row:
                #print(f"==checking for top_right: {cell.value}")
                if self.is_table_top_right(sheet, cell.row - 1, cell.column - 1):
                    table.right = cell.column - 1
                    break

            # get all tables left bottom, only check 1 column
            rows = list(sheet.iter_rows())

            target_column = []
            for m, row in enumerate(rows):
                # don't consider the rows above table.top
                if m <= table.top:
                    continue

                if m > self.max_row:
                    break
                target_column.append(row[table.left])

            #print(f"==checking for bottom_left: {target_column}")
            for cell in target_column:
                if self.is_table_bottom_left(sheet, cell.row - 1, cell.column - 1):
                    #print(f"=== found bottom, cell: #{cell.value}")
                    table.bottom = cell.row - 1
                    break

        print("=== done")


    def get_content_from_sheet(self, sheet, table):
        result = []
        rows = list(sheet.iter_rows())

        # to parse sheet2, we need to one more column to right, and one more row to down.
        for i in range(table.top, table.bottom + 2):
            temp_row = []
            for j in range(table.left, table.right + 2):
                #print(f"=== i:{i}, j:{j}, value: {rows[i][j].value}")
                temp_row.append(rows[i][j].value)

            result.append(temp_row)

        return result

    def get_csv_for_table_content(self, content):

        sio = io.StringIO()
        writer = csv.writer(sio)
        writer.writerows(content)

        csv_string = sio.getvalue()
        return csv_string

    def save_as_csv(self, filename, content):

        writer = csv.writer(open(filename, 'w'))
        for row in content:
            writer.writerow(row)

